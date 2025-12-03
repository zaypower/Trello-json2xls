import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE = "todo.json"
OUTPUT_FILE = "todo_export.xlsx"


def normalize_date(s: str | None):
    if not s:
        return ""
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return s

def safe_sheet_name(name: str) -> str:
    forbidden = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in forbidden:
        name = name.replace(ch, '_')
    return name[:31]

# Загружаем JSON
with open(INPUT_FILE, "r", encoding="utf-8") as f:
    data = json.load(f)

# Списки Trello
lists = {item["id"]: item["name"] for item in data.get("lists", [])}

# Карточки
cards_raw = data.get("cards", [])

# Группировка
grouped = {}
for c in cards_raw:
    list_name = lists.get(c["idList"], "UNKNOWN")
    grouped.setdefault(list_name, []).append({
        "shortId": c.get("idShort"),
        "name": c.get("name"),
        "dateLastActivity": normalize_date(c.get("dateLastActivity"))
    })


def autofit_columns(ws):
    """Автоматическая ширина колонок."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


# Создаём XLSX
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="DDDDDD")
row_fill = PatternFill("solid", fgColor="F7F7F7")

for list_name, items in grouped.items():
    ws = wb.create_sheet(title=safe_sheet_name(list_name))
    ws.append(["shortId", "name", "dateLastActivity"])


    # оформляем заголовок
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # строки
    for i, item in enumerate(items, start=2):
        ws.append([
            item["shortId"],
            item["name"],
            item["dateLastActivity"]
        ])

        # чередование строк
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = row_fill

    # автоширина
    autofit_columns(ws)

wb.save(OUTPUT_FILE)
print(f"Готово: {OUTPUT_FILE}")
